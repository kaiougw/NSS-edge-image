# 若直接執行本程式，程式會將子目錄中的BMP檔案重新命名並搬出資料夾，對每個BMP圖檔數值化解析，並輸出解析圖
# 需要NSS的檔案是LOT@slot@datetime@300RXM06@EDL_2，沒有_2的不適用
# 若不須重解壓縮，新命名及搬運的功能，則by pass "decompress()"這個function即可

from __future__ import annotations

import os
import shutil
import zipfile
import tempfile
import cv2
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import scipy.signal
import streamlit as st
import py7zr
import gc

# 功能: 讀取指定的excel，並將欄寬自動最佳化 (Automatically adjust column widths in Excel)
# 使用方式: 先指定name_of_wb為想要自動化的workbook檔案名稱
def auto_fit(name_of_wb=''):  # Input: Excel file name
    if name_of_wb != '':
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter

        # 讀取已經存在的特定檔案
        wb = load_workbook(name_of_wb)  # ()裡面是完整的檔案名稱
        # 對每個分頁做auto_fit
        for i in wb.sheetnames:
            ws = wb[i]
            # auto_fit:
            for letter_num in range(1, ws.max_column + 1):
                max_width = 0
                letter = get_column_letter(letter_num)  # 數字轉A, B, C.....AA, AB, AC......
                for row_number in range(1, ws.max_row + 1):
                    # 用try避免讀到空白時跳error卡住
                    try:  # 取特定欄位A1的值，要寫ws['A1'].value，依此類推
                        if len(ws[
                                   f'{letter}{row_number}'].value) > max_width:  # update max_width if the length of cell value is longer than current max_width
                            max_width = len(ws[f'{letter}{row_number}'].value)
                    except:
                        pass
                ws.column_dimensions[letter].width = (max_width + 2) * 1.2
        wb.save(name_of_wb)

def moving_average(x, w):
    """
    Input:
        x: 1D array
        w: window size (integer)
    """
    return np.convolve(x, np.ones(w), 'valid') / w

def rolling_window(a, window):
    """
    Input:
        a: 1D array
        window: window size (integer)
    """
    shape = a.shape[:-1] + (a.shape[-1] - window + 1, window)
    strides = a.strides + (a.strides[-1],)
    return np.lib.stride_tricks.as_strided(a, shape=shape, strides=strides)

def find_nonblack(npy1d, threshold):
    """
    Input:
        npy1d: 1D array
        threshold: value to compare against (integer)
    Output:
        index (int) of the first element >= threshold, or -1 if none
    """
    peaks = scipy.signal.find_peaks_cwt(npy1d, 10)
    # plt.plot(col)
    # plt.plot(peaks ,col[peaks],'o')
    # plt.show()
    res = -1
    peaks = scipy.signal.find_peaks_cwt(npy1d, 10)
    for p in peaks:
        if npy1d[p] >= threshold:
            res = p
            break
    return res

def convert_nss_rawimage(img_file):
    """
    Process a raw NSS BMP image and detect the notch position to reconstruct the full wafer edge image.
    Input:
        img_file: path to the BMP image file (string)
    Output:
        img_v1: processed image as a 2D array
    """
    # img_file='.\\P1276BM\\mjl\\1-9 (SILTRONIC-WF)_EDU_0.bmp'
    img_v = cv2.imread(img_file, cv2.IMREAD_GRAYSCALE)
    # print(img_v.shape)

    if img_v is None:
        print("Invalid Image.")
        return np.array([])

    if img_v.shape[0] == 384000 and img_v.shape[1] == 512:
        col_u = img_v[:, 5]  # detection upper notch white spot
        col_l = img_v[:, 507]  # detection lower notch white spot
        res_u = 0
        res_l = 0

        for x, val in enumerate(col_u):  # scan upper column for first pixel >=254
            if val >= 254:
                res_u = x
                break
        for x, val in enumerate(col_l):  # scan lower column for first pixel >=254
            if val >= 254:
                res_l = x
                break

        res = np.max([res_u, res_l])  # take upper or lower detected position as notch position

        if res > 0:  # if notch is found, proceed to reconstruct image and slice image into two parts
            end_1 = int((img_v.shape[0] - res) / 1038) - 2
            top_1 = (359 - end_1 + 1)
            img_1 = img_v[res:res + end_1 * 1038, :]
            img_2 = img_v[res - top_1 * 1038:res, :]

            img_v1 = cv2.vconcat([img_1, img_2])  # concatenate two parts vertically to form a continuous wafer image
            # print(img_v1.shape)
            for i in range(0, 360):
                # cv2.line(img_v1, (0,i*1038), (45,i*1038), (255, 255, 255), 3)
                # cv2.putText(img_v1, str(i), (10,i*1038+30), cv2.FONT_HERSHEY_SIMPLEX,  1, (255, 255, 255), 1, cv2.LINE_AA)

                cv2.line(img_v1, (img_v1.shape[1] - 45, i * 1038), (img_v1.shape[1], i * 1038), (255, 255, 255), 3)
                cv2.putText(img_v1, str(i), (img_v1.shape[1] - 60, i * 1038 + 30), cv2.FONT_HERSHEY_SIMPLEX, 1,
                            (255, 255, 255), 1, cv2.LINE_AA)

            img_v1 = cv2.rotate(img_v1, cv2.ROTATE_90_CLOCKWISE)
            img_v1 = cv2.flip(img_v1, 1)  # flip left-right
            # cv2.imwrite('.\\P1276BM\\mjl\\test.png', img_v1)
            return img_v1
        else:
            print("Can't find the notch.")
            return np.array([])
    else:
        print("Invalid Image.")
        return np.array([])

def turning_points(array):
    '''
    Finds the turning points within an 1D array and returns the indices of the minimum and
    maximum turning points in two separate lists.

    Input:
        array: 1D array
    Output:
        idx_min: list of indices of minimum turning points
        idx_max: list of indices of maximum turning points
    '''
    idx_max, idx_min = [], []
    if (len(array) < 3):
        return idx_min, idx_max

    NEUTRAL, RISING, FALLING = range(3)

    def get_state(a, b):
        if a < b: return RISING
        if a > b: return FALLING
        return NEUTRAL

    ps = get_state(array[0], array[1])
    begin = 1
    for i in range(2,
                   len(array)):  # if trend changes (from rising to falling or falling to rising), record turning point
        s = get_state(array[i - 1], array[i])  # compare each pair of values
        if s != NEUTRAL:
            if ps != NEUTRAL and ps != s:
                if s == FALLING:
                    idx_max.append((begin + i - 1) // 2)
                else:
                    idx_min.append((begin + i - 1) // 2)
            begin = i
            ps = s
    return idx_min, idx_max

def process_bmp0(bmpfile):
    """
    Process a BMP file to analyze wafer edge image, compute roughness, and generate related charts and CSV files.

    Input:
        bmpfile: path to the BMP image file (string)
    Output:
        List containing analysis results
    """
    # print(bmpfile)

    nss_img_path = os.path.dirname(bmpfile) + '/'
    img_file = bmpfile  # './/P1276BM//' + f + '.png'
    img_v = convert_nss_rawimage(img_file)

    if img_v.size == 0:
        print('Invalid NSS bmp file...!')
        return []

    # print(nss_img_path + os.path.basename(bmpfile)[:-4])

    if img_v.shape[1] < (360 * 1038):
        print('Invalid NSS bmp file...!')
        return []
    else:
        # detection wafer position
        col1 = img_v[:, 5000]
        col2 = img_v[:, 5500]
        col3 = img_v[:, 10000]
        col4 = img_v[:, 10500]
        col5 = img_v[:, 20000]
        col6 = img_v[:, 21000]
        th_ = 50
        res1 = find_nonblack(col1, th_)
        res2 = find_nonblack(col2, th_)
        res3 = find_nonblack(col3, th_)
        res4 = find_nonblack(col4, th_)
        res5 = find_nonblack(col5, th_)
        res6 = find_nonblack(col6, th_)
        candidates = [r for r in [res1, res2, res3, res4, res5, res6] if r >= 0]
        if not candidates:
            print('Invalid NSS bmp file...!')
            return []
        res = np.min(candidates)
        c1 = res + 10
        c2 = res + 230
        img_v0 = img_v[c1:c2, 10000:10500]
        # plt.imshow(img_v0,cmap=plt.cm.gray)
        # plt.show()
        # plt.clf()

        # column stdev
        a0 = np.std(img_v[c1:c2, :].astype(np.float32), axis=0)
        x = np.arange(0, a0.shape[0])

        # row stdev 1038
        mv_sdev_window = 1038  # 519
        a01 = np.std(rolling_window(a0, mv_sdev_window), 1) ** 2
        ax = moving_average(x / 1038, mv_sdev_window)  # 1038

        # Ra & Q95
        a02 = a01[1038 * 3:1038 * -2]
        Ra = np.sum(np.abs(a02)) / a02.shape[0]
        Q95 = np.percentile(a02, 95, axis=0)
        Q90 = np.percentile(a02, 90, axis=0)
        Q50 = np.percentile(a02, 50, axis=0)
        # rpt.append([img_file,Ra,Q50,Q90,Q95])

        # 360 ra
        rpt_360 = []
        for i in range(3, 358):
            a_ = a01[1038 * i:1038 * (i + 1)]
            ra_ = np.sum(np.abs(a_)) / a_.shape[0]
            rpt_360.append([i, ra_])

        # find peaks
        a_360_x = np.array(rpt_360)[:, 0]
        a_360_y = np.array(rpt_360)[:, 1]
        idx_min, idx_max = turning_points(a_360_y)
        rpt_360_peaks = []
        for i in idx_max:
            rpt_360_peaks.append([a_360_x[i], a_360_y[i]])
        arr = np.array(rpt_360_peaks)
        a_360_peaks = arr[arr[:, 1].argsort()]

        # plot by deg chart
        plt.figure(figsize=(8, 4))
        plt.title('NSS EDGE Image Quality Ra by Deg')
        plt.plot(a_360_x, a_360_y, label=img_file)
        plt.scatter(a_360_peaks[-10:, 0], a_360_peaks[-10:, 1], facecolors='none', edgecolors='r')
        plt.scatter(a_360_peaks[-20:-10, 0], a_360_peaks[-20:-10, 1], facecolors='none', edgecolors='g')
        plt.ylim(0, 100)
        plt.xticks(np.arange(0, 360, 10.0))
        plt.xticks(rotation=90, ha='left')
        plt.xlabel('Angle NotchDown CW')
        plt.legend()
        plt.savefig(nss_img_path + os.path.basename(bmpfile)[:-4] + '_360chart.jpg')
        # plt.show()
        plt.clf()
        plt.close()

        # excluded notch before 2500 and last 2500 pixels
        plt.figure(figsize=(8, 4))
        plt.title('NSS EDGE Image Quality Ra=' + str(round(Ra, 2)))
        plt.plot(ax[1038 * 3:1038 * -2], a01[1038 * 3:1038 * -2], label=img_file)
        plt.ylim(0, 100)
        plt.xlabel('Angle NotchDown CW')
        plt.xticks(np.arange(0, 360, 10.0))
        plt.xticks(rotation=90, ha='left')
        plt.legend()
        plt.savefig(nss_img_path + os.path.basename(bmpfile)[:-4] + '_chart.jpg')
        # plt.show()
        plt.clf()
        plt.close()
        # export to save csv files
        np.savetxt(nss_img_path + os.path.basename(bmpfile)[:-4] + '_peaks.csv', a_360_peaks, fmt='%1.3f',
                   delimiter=',')
        np.savetxt(nss_img_path + os.path.basename(bmpfile)[:-4] + '_360.csv', np.array(rpt_360), fmt='%1.3f',
                   delimiter=',')
        # np.savetxt(nss_img_path + os.path.basename(f)[:-4] + '.csv',a01, fmt='%1.3f',delimiter=',')

        # save images
        for i in np.arange(1038 * 3, a0.shape[0] - 1038 * 2):
            x0 = int(i) + 519
            x1 = int(i + 1) + 519
            y0 = int(a01[i] / 100 * 200)
            y1 = int(a01[i + 1] / 100 * 200)
            cv2.line(img_v, (x0, img_v.shape[0] - y0 - 10), (x1, img_v.shape[0] - y1 - 10), (255, 255, 255), 1)

        # print(nss_img_path + os.path.basename(f)[:-4] + '.png')
        # save whole wafer image
        cv2.imwrite(nss_img_path + os.path.basename(bmpfile)[:-4] + '.png', img_v)

        # crop image for top 20 peaks
        tmp_dir = os.path.dirname(bmpfile) + '/' + os.path.basename(bmpfile)[:-4]
        if not os.path.isdir(tmp_dir):
            os.mkdir(tmp_dir)
        for i in a_360_peaks[-20:, 0]:
            img_v0 = img_v[:, int((i - 0.1) * 1038):int((i + 2) * 1038)]
            cv2.imwrite(tmp_dir + '/' + str(int(i)) + '.png', img_v0)

        del img_v, a0, a01, a02, rpt_360, a_360_x, a_360_y
        gc.collect()

        return [img_file, Ra, Q50, Q90, Q95]

def process_bmp(bmpfile):
    """
    Process a BMP file to analyze wafer edge image, compute roughness, and generate related charts and CSV files.

    Input:
        bmpfile: path to the BMP image file (string)
    Output:
        List containing analysis results
    """
    # print(bmpfile)
    f = bmpfile
    nss_img_path = os.path.dirname(bmpfile) + '/'
    img_file = bmpfile  # './/P1276BM//' + f + '.png'
    img_v = convert_nss_rawimage(img_file)

    # early exit if conversion failed to avoid attribute access on empty arrays
    if img_v.size == 0:
        print('Invalid NSS bmp file...!')
        return []

    # print(nss_img_path + os.path.basename(bmpfile)[:-4])
    # row stdev 1038
    # mv_sdev_window_0=5 #519 short wave
    mv_sdev_window_1 = 1038  # 519 long wave

    if img_v.shape[1] < (360 * mv_sdev_window_1):
        print('Invalid NSS bmp file...!')
        return []
    else:
        # detection wafer position
        col1 = img_v[:, 5000]
        col2 = img_v[:, 5500]
        col3 = img_v[:, 10000]
        col4 = img_v[:, 10500]
        col5 = img_v[:, 20000]
        col6 = img_v[:, 21000]
        th_ = 50
        res1 = find_nonblack(col1, th_)
        res2 = find_nonblack(col2, th_)
        res3 = find_nonblack(col3, th_)
        res4 = find_nonblack(col4, th_)
        res5 = find_nonblack(col5, th_)
        res6 = find_nonblack(col6, th_)
        # guard against no valid indices
        candidates = [r for r in [res1, res2, res3, res4, res5, res6] if r >= 0]
        if not candidates:
            print('Invalid NSS bmp file...!')
            return []
        res = np.min(candidates)
        c1 = res + 10
        c2 = res + 230
        # img_v0=img_v[c1:c2,10000:10500]
        # plt.imshow(img_v0,cmap=plt.cm.gray)
        # plt.show()
        # plt.clf()

        # column stdev
        roi = img_v[c1:c2, :].astype(np.float32)  # cast to float32 to reduce memory
        a0_0 = np.max(roi, axis=0) - np.min(roi, axis=0)
        a0_1 = np.std(roi, axis=0)
        x = np.arange(0, a0_1.shape[0])

        # a01_0=a0_0-np.mean(a0_0)  #1038
        a01_0 = a0_0 - np.min(a0_0[mv_sdev_window_1 * 3:mv_sdev_window_1 * -3])  # 1038
        a01_1 = np.std(rolling_window(a0_1, mv_sdev_window_1), 1) ** 2  # 1038
        # angles
        ax_0 = x / mv_sdev_window_1  # 1038
        ax_1 = moving_average(x / mv_sdev_window_1, mv_sdev_window_1)  # 1038

        # Ra & Q95
        a02_0 = a01_0[mv_sdev_window_1 * 3:mv_sdev_window_1 * -3]
        a02_1 = a01_1[mv_sdev_window_1 * 3:mv_sdev_window_1 * -3]

        Ra_0 = np.sum(np.abs(a02_0)) / a02_0.shape[0]
        Q99_0 = np.percentile(a02_0, 99, axis=0)
        Q90_0 = np.percentile(a02_0, 90, axis=0)
        Q50_0 = np.percentile(a02_0, 50, axis=0)

        Ra_1 = np.sum(np.abs(a02_1)) / a02_1.shape[0]
        Q99_1 = np.percentile(a02_1, 99, axis=0)
        Q90_1 = np.percentile(a02_1, 90, axis=0)
        Q50_1 = np.percentile(a02_1, 50, axis=0)

        # rpt.append([img_file,Ra_0,Q50_0,Q90_0,Q99_0,Ra_1,Q50_1,Q90_1,Q99_1])

        # 360 ra
        rpt_360 = []
        for i in range(3, 358):
            a_0 = a01_0[mv_sdev_window_1 * i:mv_sdev_window_1 * (i + 1)]
            a_1 = a01_1[mv_sdev_window_1 * i:mv_sdev_window_1 * (i + 1)]

            a_0_98 = np.percentile(a_0, 98, axis=0)
            ra_0 = np.mean(a_0_98)
            ra_1 = np.sum(np.abs(a_1)) / a_1.shape[0]
            rpt_360.append([i, ra_0, ra_1])

        # find peaks
        a_360_x = np.array(rpt_360)[:, 0]
        a_360_y0 = np.array(rpt_360)[:, 1]
        a_360_y1 = np.array(rpt_360)[:, 2]
        idx_min0, idx_max0 = turning_points(a_360_y0)
        idx_min1, idx_max1 = turning_points(a_360_y1)
        rpt_360_peaks0 = []
        rpt_360_peaks1 = []

        for i in idx_max0:
            rpt_360_peaks0.append([a_360_x[i], a_360_y0[i]])
        for i in idx_max1:
            rpt_360_peaks1.append([a_360_x[i], a_360_y1[i]])

        arr0 = np.array(rpt_360_peaks0)
        arr1 = np.array(rpt_360_peaks1)
        a_360_peaks0 = arr0[arr0[:, 1].argsort()]
        a_360_peaks1 = arr1[arr1[:, 1].argsort()]

        # plot by deg chart
        plt.figure(figsize=(8, 4))
        plt.title('NSS EDGE Image Quality Ra by Deg')
        plt.plot(a_360_x, a_360_y0, label=os.path.basename(img_file) + '(raw)')
        plt.plot(a_360_x, a_360_y1, label=os.path.basename(img_file) + '(Moving Avg/Original Ra)')

        plt.scatter(a_360_peaks0[-10:, 0], a_360_peaks0[-10:, 1], facecolors='none', edgecolors='r')
        plt.scatter(a_360_peaks0[-20:-10, 0], a_360_peaks0[-20:-10, 1], facecolors='none', edgecolors='g')

        plt.scatter(a_360_peaks1[-10:, 0], a_360_peaks1[-10:, 1], facecolors='none', edgecolors='r')
        plt.scatter(a_360_peaks1[-20:-10, 0], a_360_peaks1[-20:-10, 1], facecolors='none', edgecolors='g')

        plt.ylim(0, 150)
        plt.xticks(np.arange(0, 360, 10.0))
        plt.xticks(rotation=90, ha='left')
        plt.xlabel('Angle NotchDown CW')
        plt.legend()
        plt.savefig(nss_img_path + os.path.basename(f)[:-4] + '_360chart.jpg')
        # plt.show()
        plt.clf()
        plt.close()

        # excluded notch before 2500 and last 2500 pixels
        # =================================================朱sir原始檔案中有輸出，但分析時基本上用不到，故先mark掉，要用時打開就好
        '''
        plt.figure(figsize=(8,4))
        plt.title('NSS EDGE Image Quality Ra_raw=' + str(round(Ra_0,2))+ 'Ra_mv=' + str(round(Ra_1,2)) )
        plt.plot(ax_0[mv_sdev_window_1*3:mv_sdev_window_1*-2],a01_0[mv_sdev_window_1*3:mv_sdev_window_1*-2],label=os.path.basename(img_file) + '(raw)')
        plt.plot(ax_1[mv_sdev_window_1*3:mv_sdev_window_1*-2],a01_1[mv_sdev_window_1*3:mv_sdev_window_1*-2],label=os.path.basename(img_file) + '(Moving Avg/Original Ra)')
        plt.ylim(0,150)
        plt.xlabel('Angle NotchDown CW')
        plt.xticks(np.arange(0, 360, 10.0))
        plt.xticks(rotation=90, ha='left')
        plt.legend()
        plt.savefig(nss_img_path + os.path.basename(f)[:-4] + '_chart.jpg')
        #plt.show()
        plt.clf()
        plt.close()
        '''
        # =================================================朱sir原始檔案中有輸出，但分析時基本上用不到，故先mark掉，要用時打開就好

        # 朱sir原始檔案中有輸出每片wafer的csv檔案，但分析時基本上用不到，故先mark掉，要用時打開就好
        # ===================================================================================================================
        '''
        #export to save csv files
        np.savetxt(nss_img_path + os.path.basename(f)[:-4] + '_raw_peaks.csv',a_360_peaks0, fmt='%1.3f',delimiter=',')
        np.savetxt(nss_img_path + os.path.basename(f)[:-4] + '_mv_peaks.csv',a_360_peaks1, fmt='%1.3f',delimiter=',')
        np.savetxt(nss_img_path + os.path.basename(f)[:-4] + '_360.csv',np.array(rpt_360), fmt='%1.3f',delimiter=',')
        #np.savetxt(nss_img_path + os.path.basename(f)[:-4] + '.csv',a01, fmt='%1.3f',delimiter=',')
        '''
        # ===================================================================================================================

        # save images
        for i in np.arange(mv_sdev_window_1 * 3, a0_1.shape[0] - mv_sdev_window_1 * 3):
            x0 = int(i)  # +519
            x1 = int(i + 1)  # +519
            y0 = int(a01_0[i] / 100 * 200)
            y1 = int(a01_0[i + 1] / 100 * 200)

            x2 = int(i) + int(mv_sdev_window_1 / 2)
            x3 = int(i + 1) + int(mv_sdev_window_1 / 2)
            y2 = int(a01_1[i] / 100 * 200)
            y3 = int(a01_1[i + 1] / 100 * 200)
            cv2.line(img_v, (x0, img_v.shape[0] - y0 - 10), (x1, img_v.shape[0] - y1 - 10), (255, 255, 255), 1)
            cv2.line(img_v, (x2, img_v.shape[0] - y2 - 10), (x3, img_v.shape[0] - y3 - 10), (255, 255, 255), 2)
            # print(nss_img_path + os.path.basename(f)[:-4] + '.png')
        # save whole wafer image
        cv2.imwrite(nss_img_path + os.path.basename(f)[:-4] + '.png', img_v)

        # crop image for top 20 peaks
        tmp_dir = os.path.dirname(f) + '/' + os.path.basename(f)[:-4]
        if not os.path.isdir(tmp_dir):
            os.mkdir(tmp_dir)
        for i in a_360_peaks0[-20:, 0]:
            img_v0 = img_v[:, int((i - 0.3) * mv_sdev_window_1):int((i + 1.6) * mv_sdev_window_1)]
            cv2.imwrite(tmp_dir + '/' + str(int(i)) + '.png', img_v0)
        for i in a_360_peaks1[-20:, 0]:
            if not os.path.isfile(tmp_dir + '/' + str(int(i)) + '.png'):
                img_v0 = img_v[:, int((i - 0.3) * mv_sdev_window_1):int((i + 1.6) * mv_sdev_window_1)]
                cv2.imwrite(tmp_dir + '/' + str(int(i)) + '.png', img_v0)

        # free large arrays to reduce peak memory
        del img_v, roi, a0_0, a0_1, a01_0, a01_1, a02_0, a02_1, rpt_360, a_360_x, a_360_y0, a_360_y1
        gc.collect()

        return [img_file, Ra_0, Q50_0, Q90_0, Q99_0, Ra_1, Q50_1, Q90_1, Q99_1]
        # else:
    #    print('Execption...')
    #    return []

def decompress():
    # 獲取當前目錄
    mypath = os.path.dirname(os.path.realpath(__file__))
    list_of_all_file = []
    # 遞迴列出所有7z檔案的絕對路徑
    for root, dirs, files in os.walk(mypath):
        for f in files:
            if f.find('.7z') != -1:
                fullpath = os.path.join(root, f)
                list_of_all_file.append(fullpath)

                # 依序解壓縮並開一個資料夾存BMP檔案，資料夾檔案名稱為lot+slot位置，檔案名稱為lot+slot位置
    for i in list_of_all_file:
        position = i.index('@')
        waferid = i[position - 9:position] + i[position + 1:position + 3]
        os.mkdir(waferid)  # 建立資料夾
        archive = py7zr.SevenZipFile(i, mode='r')  # 讀取7z檔案
        archive.extractall(path=mypath + '/' + waferid)  # 把讀到的檔案存在新建的資料夾中
        archive.close()
        list_of_all_bmp = []
        # 列出資料夾內bmp檔案的絕對路徑
        for root, dirs, files in os.walk(mypath + '/' + waferid):
            for f in files:
                if f.find('.bmp') != -1:
                    fullpath = os.path.join(root, f)
                    list_of_all_bmp.append(fullpath)
        # 重新命名bmp檔案
        for j in list_of_all_bmp:
            old_file = j
            new_file = os.path.join(os.path.dirname(j), waferid + '.bmp')
            os.rename(old_file, new_file)
    # 再次遞迴列出所有bmp檔案的絕對路徑
    list_of_all_file = []
    for root, dirs, files in os.walk(mypath):
        for f in files:
            if f.find('.bmp') != -1:
                fullpath = os.path.join(root, f)
                list_of_all_file.append(fullpath)
    # 移動所有bmp檔案
    for i in list_of_all_file:
        file_source = i
        file_destination = mypath
        shutil.move(file_source, file_destination)

def rename_move_rawdata():
    # 獲取當前目錄
    mypath = os.path.dirname(os.path.realpath(__file__))
    list_of_all_file = []
    # 遞迴列出所有bmp檔案的絕對路徑
    for root, dirs, files in os.walk(mypath):
        for f in files:
            if f.find('.bmp') != -1:
                fullpath = os.path.join(root, f)
                list_of_all_file.append(fullpath)
    # 重新命名所有bmp檔案
    for i in list_of_all_file:
        position = i.index('@')
        waferid = i[position - 9:position] + i[position + 1:position + 3]
        old_file = i
        new_file = os.path.join(os.path.dirname(i), waferid + '.bmp')
        os.rename(old_file, new_file)
    # 再次遞迴列出所有bmp檔案的絕對路徑
    list_of_all_file = []
    for root, dirs, files in os.walk(mypath):
        for f in files:
            if f.find('.bmp') != -1:
                fullpath = os.path.join(root, f)
                list_of_all_file.append(fullpath)
    # 移動所有bmp檔案
    for i in list_of_all_file:
        file_source = i
        file_destination = mypath
        shutil.move(file_source, file_destination)


st.set_page_config(page_title="NSS Edge Image", layout="wide")
st.title("NSS Edge Image")

st.caption("Enter a file path or browse below")

network = [f"{d}:/".replace("/", os.sep) for d in ["H", "M", "Z"] if os.path.exists(f"{d}:/".replace("/", os.sep))]  # network drives limited to H:, M:, Z:

typed_path = st.text_input("Enter a file path", value=st.session_state.get("typed_path", ""), label_visibility="collapsed", placeholder="Enter a file path", key="typed_path")
path_override = None
if isinstance(typed_path, str) and typed_path.strip():
    tp = typed_path.strip().replace("/", os.sep) # replace / with \
    tp = os.path.normpath(tp) # remove redundant .. or / and normalize path

    drive, _ = os.path.splitdrive(tp) # extract drive
    if drive:
        drive_root = os.path.join(drive + ":", "")
        if os.path.exists(drive_root):
            st.session_state["selected_network"] = drive_root # auto-select the drive in `selected_network` selectbox based on user input

    try: # check if path include "EDL" or "EDU" as first folder, then auto-select
        after_drive = tp[len(drive + os.sep):] if drive else tp
        folder1 = after_drive.split(os.sep)[0] if after_drive else ""
        if folder1 in ("EDL", "EDU"):
            st.session_state["selected_folder"] = folder1
    except Exception:
        pass

    if os.path.isdir(tp):
        path_override = tp

path_selection = st.empty()
selected_network = st.selectbox("Select a file path", network, key="selected_network", label_visibility="hidden", index=None, placeholder="Select a network drive")
if not isinstance(selected_network, str):
    st.stop()
ROOT = selected_network

path_selection.markdown(f"`{ROOT}`")

selected_folder = False
if isinstance(selected_network, str) and selected_network.startswith(("M", "Z")): # M: and Z: drives contain raw data
    selected_folder = st.selectbox("Select a bevel folder", ["EDL", "EDU"], key="selected_folder", label_visibility="collapsed", index=None, placeholder="Select a folder")
    if not isinstance(selected_folder, str):
        st.stop()
    ROOT = os.path.join(selected_network, selected_folder)

typed_steps = [] # list of sequence of folder names
typed_zip = None # .zip file name if user typed the full file path
typed_ptr = 0 # index to keep track of how many files have been auto-navigated through

if path_override:
    if tp.lower().endswith(".zip") and os.path.isfile(tp): # if typed path ends with .zip and points to an existing file
        typed_zip = os.path.basename(tp)  # .zip file name = the base name of the typed file path
        path_override = os.path.dirname(tp) #  path_override = parent folder (i.e., file path leading to .zip file)
    try:
        rel = os.path.relpath(path_override, ROOT) if ROOT else None # rel = relative path from ROOT (e.g., M:\EDL) to path_override (e.g., M:\EDL\2025-08\C\4300) -> rel becomes 2025-08\C\4300
        if rel and not rel.startswith(os.pardir): # if rel does not start with ".." meaning that it is inside ROOT
            typed_steps = [p for p in rel.split(os.sep) if p] # then split the path into folder components using os.sep (/ or \)
        else: # if rel starts with ".." meaning that it is outside ROOT
            typed_steps = [p for p in os.path.normpath(path_override).split(os.sep) if p] # then normalize and split the path
    except Exception:
        typed_steps = [p for p in os.path.normpath(path_override).split(os.sep) if p]

path_selection.markdown(f"`{ROOT}`")

selected_zip = False # for storing full path of the selected .zip file
if isinstance(ROOT, str):
    current_path = path_override or ROOT # current path starts from the path user typed; otherwise, start from ROOT
    level = 0 # index to keep track of how many files have been auto-navigated through
    while True: # Keep looping and go down each folder until a .zip file is found
        try:
            with os.scandir(current_path) as it:
                dirs = sorted([e.name for e in it if e.is_dir()])

                path_selection.markdown(f"`{current_path}`")

        except FileNotFoundError:
            st.error("Path not found.")
            st.stop()

        with os.scandir(current_path) as it:
            zips = sorted([e.name for e in it if e.is_file() and e.name.lower().endswith(".zip")])

        if zips:
            if typed_zip and typed_zip in zips:
                # Pre-fill selectbox value so the UI reflects the typed choice
                st.session_state[f"zip_select_{current_path}"] = typed_zip
                selected_zip_name = typed_zip
            else:
                selected_zip_name = st.selectbox(
                    "Select a **.zip** file. The .bmp file will be extracted.",
                    zips,
                    key=f"zip_select_{current_path}",
                    index=None,
                    placeholder="Select a .zip file"
                )
                if not isinstance(selected_zip_name, str):
                    st.stop()

            selected_zip = os.path.join(current_path, selected_zip_name)
            break  # done navigating; proceed to processing

        if len(dirs) == 0: # if no folders or .zip files are found
            st.error(f"No .zip files found under: {current_path}")
            st.stop()

        if len(dirs) == 1: # "auto-skip" folder; if it only contains one folder, automatically proceed to the next folder so that user does not need to select the single folder
            only = dirs[0]
            current_path = os.path.join(current_path, only)
        else:
            auto_choice = None
            if typed_ptr < len(typed_steps):
                candidate = typed_steps[typed_ptr]
                if candidate in dirs:
                    st.session_state[f"dir_select_level_{level}"] = candidate
                    auto_choice = candidate
                    typed_ptr += 1  # advance only when we used this step

            if auto_choice is None:
                selected_dir = st.selectbox("Select a folder", dirs, key=f"dir_select_level_{level}", label_visibility="collapsed", index=None, placeholder="Select a folder")
                if not isinstance(selected_dir, str):
                    st.stop()
            else:
                selected_dir = auto_choice

            current_path = os.path.join(current_path, selected_dir)
            level += 1

    if selected_zip:
        if st.button("Process", type="primary", key="process_selected_zip"):
            with tempfile.TemporaryDirectory(prefix="nss_zip_") as workdir:
                outdir = os.path.join(workdir, "outputs")
                os.makedirs(outdir, exist_ok=True)

                bmp_paths = []
                with zipfile.ZipFile(selected_zip) as zf:
                    for info in zf.infolist():
                        if info.filename.lower().endswith(".bmp"):
                            zf.extract(info, workdir)
                            bmp_paths.append(os.path.join(workdir, info.filename))

                if not bmp_paths:
                    st.error("No .bmp files found inside the .zip file.")
                    st.stop()

                rows = []
                for bmp in bmp_paths:
                    try:
                        row = process_bmp(bmp)
                        if row:
                            rows.append(row)
                    except Exception as e:
                        st.warning(f"Failed on {os.path.basename(bmp)}: {e}")
                    finally:
                        gc.collect()

                if not rows:
                    st.error("Processing finished but no results were produced.")
                    st.stop()

                cols = ['filename', 'Ra_raw', 'RawQ50', 'RawQ90', 'RawQ99', 'Ra_mv', 'MvQ50', 'MvQ90', 'MvQ99']
                df = pd.DataFrame(rows, columns=cols)

                summary_xlsx = os.path.join(outdir, "nss_image_summary.xlsx")
                df.to_excel(summary_xlsx, index=False)
                auto_fit(summary_xlsx)

                st.subheader("Summary")
                st.dataframe(df, use_container_width=True)
                st.download_button(
                    "Download",
                    data=open(summary_xlsx, "rb").read(),
                    file_name="nss_image_summary.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

                st.success("Done.")
